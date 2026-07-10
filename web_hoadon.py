import streamlit as st
import io
import openpyxl
import re
import os
import unicodedata
import tempfile
import urllib.request
import ssl
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# --- CẤU HÌNH GIAO DIỆN ---
st.set_page_config(page_title="Green Chicken - Kênh MT", page_icon="🍗", layout="centered")
st.markdown("""<style>.stButton>button {background-color: #2E8B57 !important; color: white !important; font-weight: bold;} h1 {color: #2E8B57 !important;}</style>""", unsafe_allow_html=True)

# --- LOGO & THÔNG TIN ---
st.title("CÔNG CỤ HỖ TRỢ KÊNH MT PRO")
st.markdown("""
### Hỗ trợ note địa chỉ vào hoá đơn 
* 📞 **Liên hệ hỗ trợ:** [0326.019.777](tel:0326019777)
* 🏭 **Email:** Torres.nam@deheus.vn
""")
st.divider()

# --- 1. CHUẨN HÓA UNICODE ---
def chuan_hoa_unicode(text):
    if not text: return ""
    return unicodedata.normalize('NFC', str(text)).replace('\xa0', ' ').strip()

# --- 2. NẠP FONT TIẾNG VIỆT (VƯỢT TƯỜNG LỬA BẢO MẬT CLOUD) ---
@st.cache_resource
def load_vietnamese_font():
    # Ưu tiên 1: Tải font bằng urllib với SSL Bypass (Vượt tường lửa Streamlit Cloud)
    font_path = os.path.join(tempfile.gettempdir(), "Roboto-Bold.ttf")
    if not os.path.exists(font_path):
        try:
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            url = "https://raw.githubusercontent.com/google/fonts/main/ofl/roboto/Roboto-Bold.ttf"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, context=ctx, timeout=10) as response, open(font_path, 'wb') as out_file:
                out_file.write(response.read())
        except: pass
        
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont("FontTiengViet", font_path))
            return "FontTiengViet"
        except: pass

    # Ưu tiên 2: Tìm font anh Nam tự upload lên Github (Nếu có)
    local_fonts = ["Roboto-Bold.ttf", "Arial.ttf", "arialbd.ttf", "Arial Bold.ttf"]
    for lf in local_fonts:
        if os.path.exists(lf):
            try:
                pdfmetrics.registerFont(TTFont("FontTiengViet", lf))
                return "FontTiengViet"
            except: pass

    # Ưu tiên 3: Dùng font hệ thống nếu đang chạy trên máy Mac/Win
    system_fonts = [
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        "/Library/Fonts/Arial Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf"
    ]
    for path in system_fonts:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("FontTiengViet", path))
                return "FontTiengViet"
            except: continue

    # Nếu hoàn toàn thất bại: Báo lỗi chứ KIÊN QUYẾT KHÔNG DÙNG FONT TIẾNG ANH
    return None

# --- 3. LẤY MÃ SO CHÍNH XÁC ---
def lay_ma_so_soan_hang(text_value):
    if not text_value: return None
    text = chuan_hoa_unicode(text_value)
    digits = "".join(re.findall(r'\d+', text))
    return digits[-6:] if len(digits) >= 6 else (digits if digits else None)

# --- 4. BỘ LỌC ĐỊA CHỈ SIÊU THỊ ---
def loc_ten_sieu_thi_pro(raw_note):
    text = chuan_hoa_unicode(raw_note)
    if not text: return ""
    
    typo_map = {"JJIMART": "FUJIMART", "FUJI ": "FUJIMART ", "WINMAT": "WINMART", "DELI ": "DELICA ", "THANH DO": "THÀNH ĐÔ", "BRG ": "BRG "}
    for wrong, right in typo_map.items():
        text = re.sub(re.escape(wrong), right, text, flags=re.IGNORECASE)
        
    text = re.sub(r'\d+h\d*(-\d+h\d*)?|\d+:\d+|trước\s*\d+h', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'Giao\s*xe\s*máy|Giao\s*xe|Giao\s*hàng|Giao|Xuất\s*kho|Xuất\s*cho|Xuất', ' ', text, flags=re.IGNORECASE)
    
    cac_cum = [c.strip() for c in re.split(r'[,;]', text) if c.strip()]
    if not cac_cum: return ""
    
    brands = ["FUJIMART", "BRG", "DELICA", "THÀNH ĐÔ", "INTRACOM", "WINMART", "AEON", "LOTTE"]
    
    for cum in cac_cum:
        cum_upper = cum.upper()
        for b in brands:
            if b in cum_upper:
                clean_loc = re.sub(r'\d+', ' ', cum_upper).replace(b, '').strip()
                clean_loc = re.sub(r'[\-\(\)\.]', ' ', clean_loc)
                clean_loc = re.sub(r'\s+', ' ', clean_loc).strip()
                
                words = clean_loc.split()
                seen, unique = set(), []
                for w in words:
                    if w not in seen:
                        unique.append(w)
                        seen.add(w)
                loc_final = " ".join(unique)
                
                if b == "THÀNH ĐÔ":
                    return f"THÀNH ĐÔ - {loc_final}" if loc_final else "THÀNH ĐÔ"
                else:
                    return f"{b} {loc_final}".strip() if loc_final else b
                    
    for cum in cac_cum:
        cum_clean = re.sub(r'\s+', ' ', cum).strip()
        if len(cum_clean) > 4 and not any(q in cum_clean.upper() for q in ["TP. HÀ NỘI", "HÀ NỘI", "VIỆT NAM", "Q. ĐỐNG ĐA"]):
            return cum_clean.upper()
            
    return cac_cum[0].upper().strip()

# --- 5. GIAO DIỆN TẢI FILE ---
st.markdown("### 1️⃣ Tải dữ liệu lên (Hỗ trợ gộp nhiều file Excel)")
col1, col2 = st.columns(2)
with col1:
    excel_files = st.file_uploader("📊 Chọn các file Excel (.xlsx)", type=["xlsx"], accept_multiple_files=True)
with col2:
    pdf_files = st.file_uploader("📄 Chọn các file Hóa Đơn PDF (.pdf)", type=["pdf"], accept_multiple_files=True)

st.markdown("---")

# --- 6. XỬ LÝ DỮ LIỆU ---
if st.button("🚀 Bấm Để Xử Lý Dữ Liệu ", use_container_width=True, type="primary"):
    if not excel_files or not pdf_files:
        st.error("⚠️ Vui lòng tải lên ít nhất 1 file Excel và 1 file Hóa Đơn PDF!")
    else:
        # CHỐT CHẶN FONT: Kiểm tra font ngay trước khi chạy
        ten_font = load_vietnamese_font()
        if not ten_font:
            st.error("🚨 LỖI MÁY CHỦ CLOUD: Tường lửa đang chặn tải Font Tiếng Việt!")
            st.warning("🛠 **CÁCH KHẮC PHỤC DỨT ĐIỂM 100%:**\n\nAnh Nam hãy copy file font **Arial.ttf** (hoặc Roboto.ttf) từ máy tính của anh, rồi Upload thẳng lên **cùng chỗ** với file `web_hoadon.py` trên GitHub. Hệ thống sẽ tự động lấy file font anh đăng lên để dùng!")
            st.stop() # Dừng hẳn, không cho in PDF lỗi
            
        with st.spinner("⏳ Đang xử lý gộp Excel và đóng dấu PDF..."):
            try:
                so_mapping = {}
                total_rows = 0
                
                for exc_file in excel_files:
                    wb = openpyxl.load_workbook(io.BytesIO(exc_file.read()))
                    sheet = wb.active
                    header_row = [chuan_hoa_unicode(cell).upper() if cell is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                    
                    note_col_idx = -1
                    for idx, h_upper in enumerate(header_row):
                        if any(k in h_upper for k in ["NOTE", "GIAO", "ĐỊA CHỈ", "STORE"]):
                            note_col_idx = idx
                            break
                    if note_col_idx == -1: note_col_idx = 2 if len(header_row) >= 3 else 1
                    
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0] is not None and len(row) > note_col_idx and row[note_col_idx] is not None:
                            total_rows += 1
                            so_number = lay_ma_so_soan_hang(row[0])
                            clean_store_name = loc_ten_sieu_thi_pro(row[note_col_idx])
                            if so_number and clean_store_name:
                                so_mapping[so_number] = clean_store_name

                final_writer = PdfWriter()
                stamped_count = 0
                
                for pdf_file in pdf_files:
                    reader = PdfReader(io.BytesIO(pdf_file.read()))
                    for page in reader.pages:
                        page_text = chuan_hoa_unicode(page.extract_text() or "")
                        all_digits = re.findall(r'\d+', page_text)
                        
                        matched_store = None
                        for num_str in all_digits:
                            if len(num_str) >= 6:
                                ma_so_pdf = num_str[-6:]
                                if ma_so_pdf in so_mapping:
                                    matched_store = so_mapping[ma_so_pdf]
                                    break
                        
                        if matched_store:
                            stamped_count += 1
                            mediabox = page.mediabox
                            width = float(mediabox.width)
                            height = float(mediabox.height)
                            
                            packet = io.BytesIO()
                            can = canvas.Canvas(packet, pagesize=(width, height))
                            can.setFillColorRGB(1, 0, 0)
                            can.setFont(ten_font, 16) 
                            can.drawCentredString(width / 2.0, height - 25, matched_store)
                            can.save()
                            
                            packet.seek(0)
                            new_pdf = PdfReader(packet)
                            page.merge_page(new_pdf.pages[0])
                            
                        final_writer.add_page(page)
                
                output_pdf_stream = io.BytesIO()
                final_writer.write(output_pdf_stream)
                output_pdf_stream.seek(0)
                
                st.success(f"🎉 HOÀN TẤT XỬ LÝ! {len(excel_files)} file Excel ({total_rows} dòng). Đóng dấu thành công {stamped_count} trang hóa đơn!")
                st.download_button(
                    label="📥 BẤM VÀO ĐÂY ĐỂ TẢI HÓA ĐƠN HOÀN CHỈNH VỀ MÁY",
                    data=output_pdf_stream,
                    file_name="TAT_CA_HOA_DON_DE_IN.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"⚠️ Có lỗi kỹ thuật xảy ra: {str(e)}")